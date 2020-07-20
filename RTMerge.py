""" A tool for merging assay data in the form of output from AUTOPAMPA and sequencing data in the form of output from CycLS such that compound identities are associated with assay data. Assay data retention times are more accurate than a tandem MS-based retention time estimate from CycLS, so assay data is used as the anchor for comparison within a time precision window. RTMerge also calculates some molecular properties and adds them as additional columns. Note that if you merge in multiple of the same assay type (corresponding to those from AUTOPAMPA) with identically identified experiment/target mass combos, one will get clobbered.
Input: CycLS output for each sequencing run involved, AUTOPAMPA *_Results.xlsx file for each assay, Config excel sheet, amino acid definitions text file (as for CycLS).
Output: Merged data set excel sheet.

Note on additional statistics: The isotope handling as implemented will not be accurate except with deuterium and SMILES strings with isotopes in them will not be interpreted correctly.
"""
#Copywrite Chad Townsend, 2020
#Reachable at cetownse@ucsc.edu

import openpyxl,collections,itertools,sys,argparse,os,copy
from openpyxl.styles import Font
from openpyxl.cell.cell import WriteOnlyCell
from rdkit import Chem
from rdkit.Chem import Crippen

def parse_args():
    """Takes arguments from the command line and checks their validity. 
    Handles help documentation.
    """
    parser = argparse.ArgumentParser (description=__doc__)
    parser.add_argument ('config',
            help='The path to the configuration file. The configuration file is an excel file containing the majority of this script\'s parameters.')
    parser.add_argument ('-o', '--out', dest='outfile', default='',
            help='Sets the prefix of the output file. Defaults to \'Merged.xlsx\'.')
    options = parser.parse_args()
    if not options.config.endswith('.xlsx'):
        print('Error: Config file format is incorrect. Should be \'.xlsx\'.')
        sys.exit(1)
    return options
      
def parse_config (configfile):
    """Takes a properly formatted excel file and grabs all the job specs.
    Errors on missing required parameters or parameters of the wrong data types. 
    Checks that files are present.
    """
    #field name:[required True/False,A set with assay data? True/False],expected data type
    expectedfields={'Library Constraint':[True,False,'str'],'Cyclic Library? (True/False)':[True,False,'bool'],'Amino Acid Database File':[True,False,'file'], 'Mass Precision (m/z)':[True,False,'float'],'Time Precision (s)':[True,False,'float']}
    configwb = openpyxl.load_workbook(configfile, data_only=True, read_only=True)
    sheets_expected = ["Global Parameters", "Assay Data", "Experiments"]
    for name in sheets_expected:
        if name not in configwb.sheetnames:
            print('Error: Job configuration spreadsheet not properly formatted. Missing worksheet {}.'.format(name))
            sys.exit(1)
    paramsheet = configwb['Global Parameters']
    paramdict = {}
    seterror = False
    fatal = False
    for row in paramsheet.rows:
        rowlist = [x.value for x in row]
        paramdict[str(rowlist[0])]=[x for x in rowlist[1:] if x != None]
    for f in expectedfields.keys():
        if expectedfields[f][0]:
            if paramdict[f] == []:
                print('Error: Field \'{}\' is required.'.format(f))
                fatal=True
        if expectedfields[f][2] == 'str':
            for i,val in enumerate(paramdict[f]):
                try:
                    paramdict[f][i] = str(val)
                except ValueError:
                    print('Error: Field \'{}\' expects a string value.'.format(f))
                    fatal=True
        elif expectedfields[f][2] == 'int':
            for i,val in enumerate(paramdict[f]):
                try:
                    paramdict[f][i] = int(val)
                except ValueError:
                    print('Error: Field \'{}\' expects an integer value.'.format(f))
                    fatal=True
        elif expectedfields[f][2] == 'float':
            for i,val in enumerate(paramdict[f]):
                try:
                    paramdict[f][i] = float(val)
                except ValueError:
                    print('Error: Field \'{}\' expects a numerical value.'.format(f))
                    fatal=True
        elif expectedfields[f][2] == 'bool':
            for i,val in enumerate(paramdict[f]):
                if paramdict[f][i] == True or paramdict[f][i] == False:
                    pass
                elif paramdict[f][i] == 'True':
                    paramdict[f][i] = True
                elif paramdict[f][i] == 'False':
                    paramdict[f][i] = False
                else:
                    print('Error: Field \'{}\' expects a boolean value.'.format(f))
                    fatal=True
        elif expectedfields[f][2] == 'file':
            for i,val in enumerate(paramdict[f]):
                try:
                    paramdict[f][i] = str(val)
                    if not os.path.isfile(paramdict[f][i]):
                        print('Error: File {} not found.'.format(paramdict[f][i]))
                        fatal=True
                except ValueError:
                    print('Error: Field \'{}\' expects a string value.'.format(f))
                    fatal=True
        if not expectedfields[f][1]:
            paramdict[f] = paramdict[f][0]
    if fatal:
        sys.exit(1)
    #
    #Assay data file acquisition
    #
    assaysheet = configwb['Assay Data']
    paramdict['Assay Data Files'] = []
    paramdict['Assay Types'] = []
    first = True
    for row in assaysheet.rows:
        if first:
            first=False
            continue
        rowlist = [x.value for x in row]
        paramdict['Assay Data Files'].append(str(rowlist[0]))
        paramdict['Assay Types'].append(str(rowlist[1]))
    if paramdict['Assay Data Files'] == [] or paramdict['Assay Types'] == []:
        print('Error: No assay data files/types given.') #fill out later.
        sys.exit(1)
    elif len(paramdict['Assay Data Files']) != len(paramdict['Assay Types']):
        print('Error: Not all assays asigned types or type given without assay file name.')
        sys.exit(1)
    for i,val in enumerate(paramdict['Assay Types']):
        if val != 'PAMPA' and val != 'Ratio' and val != 'Integrate':
            print('Error: Invalid assay type. Valid types are PAMPA, Ratio, Integrate.')
            sys.exit(1)
    #
    #Experimental parameters.
    #
    exptsheet = configwb['Experiments']
    first = True
    paramdict['Experiment Names'] = []
    paramdict['Sequence Data File Names'] = []
    paramdict['Assay Data Rt Offset'] = []
    paramdict['Assay Data Mass Offset'] = []
    for row in exptsheet.rows:
        if first:
            first=False
            continue
        rowlist = [x.value for x in row]
        paramdict['Experiment Names'].append(str(rowlist[0]))
        paramdict['Sequence Data File Names'].append(str(rowlist[1]))
        try:
            paramdict['Assay Data Rt Offset'].append(float(rowlist[2]))
            paramdict['Assay Data Mass Offset'].append(float(rowlist[3]))
        except ValueError:
            print('Error: Non-numeric value given for experiment retention time or mass offset.')
            sys.exit(1)
    return paramdict

def parse_constraint(aadict, constraint):
    """Find the amino acid set used in the library, peptide length; calculate stereochemistry, N-alkylation, and cyclization points.
    Stereochemical estimation will only be true for amino acid SMILES strings arranged N-terminus to C-terminus and with streochemistry indicated on the alpha carbon by explicit hyrdogen declarations.
    Example residue SMILES template: N[C@@H](R)C(=O)O
    This information is also used to generate SMILES strings for each compound later.
    """
    aa_prop_dict = collections.defaultdict(lambda :{'Stereo':None,'N-alkyl':None,'cyc_points':collections.defaultdict(list)})
    constraint_positions = constraint.split(';')
    peplength = len(constraint_positions)
    constraint_positions = [x.split(',') for x in constraint_positions]
    used = set()
    for i in constraint_positions:
        for j in i:
            used.add(j)
    for aa in used:
        if '*' in aa:
            aa = aa.replace('*','')
        if aa not in aadict.keys():
            print('Error: Invalid constraint string; most likely a bad residue name.')
            sys.exit(1)
        inbracket=False
        firstbracket=True
        atcount=0
        for i,character in enumerate(aadict[aa]):
            if character == '[':
                inbracket=True
            elif character == ']':
                inbracket=False
                firstbracket=False
            elif not inbracket:
                if character.isdigit():
                    aa_prop_dict[aa]['cyc_points'][character].append(i)
            if inbracket and firstbracket:
                if character == '@':
                    atcount+=1
        if aadict[aa][1] == '(':
            aa_prop_dict[aa]['N-alkyl']=True
        else:
            aa_prop_dict[aa]['N-alkyl']=False
        if atcount == 0:
            aa_prop_dict[aa]['Stereo']='?'
        elif atcount == 1:
            aa_prop_dict[aa]['Stereo']='D'
        elif atcount == 2:
            aa_prop_dict[aa]['Stereo']='L'
    return peplength, aa_prop_dict

def parse_aadict(file_path):
    """Read an amino acid definition text file into a dictionary. Reverse of '.smi' format (name\tsmiles).
    SMILES strings are in an N->C orientation, which is required.
    """
    aadict = dict()
    keylist = []
    with open('aadatabase.txt','r') as aadb:
        for line in aadb:
            linelist = line.strip().split()
            if len(linelist) == 3:
                aadict[linelist[0]] = linelist[2].split(',')
            elif len(linelist) == 2:
                aadict[linelist[0]]=linelist[1]
            else:
                print("Error: Database formatting is incorrect.")
    return aadict
    
def parse_assaydata (paramdict):
    """Collect Assay Data from assay excel sheets relying on expected output formatting from AUTOPAMPA.py
    """
    #Expected universal for assay data: ['Experiment','Name','Mass','Peak Number','Rt (s)']
    #Only Peak Number should be unique to each assay on the same set of compounds.
    #Assays will be merged if the same type so if there are duplicate experiments, something will get clobbered.
    assaydata = collections.defaultdict(lambda : collections.defaultdict(lambda : collections.defaultdict(lambda : {})))
    assaycols = {}
    names = collections.defaultdict(lambda : collections.defaultdict(lambda : collections.defaultdict(lambda : {})))
    types_seen = set()
    for i,assay in enumerate(paramdict['Assay Data Files']):
        awb = openpyxl.load_workbook(assay)
        assay_type = paramdict['Assay Types'][i]
        if assay_type not in types_seen:
            types_seen.add(assay_type)
            rec_cols = True
        else:
            rec_cols = False
        firstrow = True
        #assumes data sheet is named by assay type or that there is only one option.
        aws = awb.active if len(awb.sheetnames) < 2 else awb[assay_type] 
        for row in aws.rows:
            rowlist = [x.value for x in row]
            if firstrow:
                firstrow=False
                if rec_cols: #assumes all of the same assay type have the same columns
                    assaycols[assay_type]=['Peak Number']
                    assaycols[assay_type].extend(rowlist[5:])
            else:
                #   assaytype    Expt              Mass            Rt          the rest
                assaydata[assay_type][rowlist[0]][round(rowlist[2],10)][rowlist[4]]=[rowlist[3]]
                assaydata[assay_type][rowlist[0]][round(rowlist[2],10)][rowlist[4]].extend(rowlist[5:])
                names[assay_type][rowlist[0]][round(rowlist[2],10)][rowlist[4]]=rowlist[1]
    return assaydata,assaycols,names

def parse_sequencedata (paramdict,assaydata):
    """Collect sequencing data from appropriate excel sheets (directed by assay data excel sheets) with formatting expectations based on CycLS output.
    """
    #Expected for sequencing data: ['Mass','Time (s)','Scan Range','Top Sequence','Score','Next Best Score','Number of Hits','Average Score','Sequencing Confidence']
    #The first three columns can safely not be output, as once matching has occured, they are no longer relevant.
    #Scan range is not carried over as it is of limited use outside of in-depth debugging of CycLS.
    seqcols = []
    seqdata = collections.defaultdict(lambda : collections.defaultdict(lambda : {}))
    for i,expt in enumerate(paramdict['Experiment Names']):
        try:
            swb = openpyxl.load_workbook(paramdict['Sequence Data File Names'][i],read_only=True)
        except IOError:
            print('Error: File {}{} not found.'.format(expt,paramdict['Sequence Data File Names'][i]))
            sys.exit(1)
        sws = swb.active
        header=True
        for row in sws.rows:
            rowvals = [x.value for x in row]
            if header:
                for i,val in enumerate(rowvals):
                    if i > 2:
                        if val not in seqcols:
                            seqcols.append(val)
                header=False
            else:
                seqdata[expt][rowvals[0]][rowvals[1]]=rowvals[3:]
    return seqdata,seqcols

def quickintervalcompare(masterlist,comparelist,interval,offset,multimatch=True):
    """Generalized generator algorithm for interval comparisons of two lists of numbers.
       Quick for data which is pre-sorted or partially pre-sorted by nature.
    """
    sml = sorted(masterlist)
    scl = sorted(comparelist)
    cind = 0
    clen = len(scl)
    if clen == 0:
        return
    for mval in sml:
        while scl[cind] + interval < mval + offset: #cval is more than 'interval' lower than mval, so increase cval.
            if cind + 1 < clen:
                cind += 1
            else:
                break
        if scl[cind] - interval > mval + offset: #cval is more than 'interval' higher than mval, so increase mval.
            continue
        if scl[cind] + interval > mval + offset: #cval is within 'interval' of mval. Do a thing.
            yield mval,scl[cind]
            if multimatch: #The first match may not be the only one within the interval.
                i = 1
                while True:
                    if cind + i >= clen:
                        break
                    if scl[cind + i] + interval > mval + offset > scl[cind + i] - interval:
                        yield mval, scl[cind + i]
                        i += 1
                    else:
                        break

def make_SMILES (compound, aadict, aa_prop_dict, cyclic):
    """Turns the list of amino acid names into a compound
    name and a SMILES string by appending N-to-C residue 
    SMILES strings together while minding the cyclization count.
    """
    name=""
    smile=""
    isfirst=True
    i = 1
    nums='1'
    #Here we properly number all the cycle breaks in the full smiles string to prevent numbering conflicts.
    for aa in compound:
        if isfirst:
            name = aa
            try:
                smile = aadict[aa]
            except KeyError:
                print('Ack! {} isn\'t in the amino acid database/dictionary'.format(aa))
                sys.exit(1)
            for problem in aa_prop_dict[aa]['cyc_points']:
                i+=1
                if i > 9:
                    nums = '%{}'.format(i)
                else:
                    nums = str(i)
            isfirst=False
        else:
            name = name + ',' + aa
            resolvedsmiles = aadict[aa]
            offset=0
            for problem in aa_prop_dict[aa]['cyc_points']:
                resolvedsmiles = resolvedsmiles[:aa_prop_dict[aa]['cyc_points'][problem][0]+offset] + nums + resolvedsmiles[aa_prop_dict[aa]['cyc_points'][problem][0]+1+offset:aa_prop_dict[aa]['cyc_points'][problem][1]+offset] + nums + resolvedsmiles[aa_prop_dict[aa]['cyc_points'][problem][1]+1+offset:]
                if len(nums) > 1:
                    offset+=2
                i += 1
                if i > 9:
                    nums = '%{}'.format(i)
                else:
                    nums = str(i)
            #Take out the Oxygen which becomes water when peptide bond is formed.
            smile = smile[:-1] + resolvedsmiles
    #Cyclize
    if cyclic:
        for x,character in enumerate(smile):
            if character == 'N':
                #Take out the last Oxygen which becomes water when peptide bond is formed.
                smile = smile[:x+1] + nums + smile[x+1:-1]
                break
        for x in range(len(smile)-1,-1,-1):
            if smile[x] == 'C':
                smile = smile[:x+1] + nums + smile[x+2:-1]
                break
    return name,smile

def fill_cell(contents, cell):
    """Spreadsheet output aid.
    """
    cell.value = contents
    return cell

def main():
    #inputs
    options = parse_args()
    paramdict = parse_config(options.config)
    aadict = parse_aadict(paramdict['Amino Acid Database File'])
    peplength, aa_prop_dict = parse_constraint(aadict,paramdict['Library Constraint'])
    SMILESparams = aadict, aa_prop_dict, paramdict['Cyclic Library? (True/False)']
    assaydata,assaycols,names = parse_assaydata(paramdict)
    seqdata,seqcols = parse_sequencedata(paramdict,assaydata)
    #Set up output spreadsheet its header columns
    mwb = openpyxl.Workbook(write_only=True)
    mws = mwb.create_sheet()
    cell_methyl = WriteOnlyCell(mws)
    cell_stereo = WriteOnlyCell(mws)
    cell_methyl.font = Font(name = 'Courier New')
    cell_stereo.font = Font(name = 'Courier New')
    cell_link = WriteOnlyCell(mws)
    cell_link.font = Font(color = '0000FF', underline='single')
    headerrow = [paramdict['Assay Types'][0]]*4
    cols = ['Expt','Name','Mass', 'Rt (s)']
    headerrow.extend(['Sequencing']*len(seqcols))
    cols.extend(seqcols)
    headerrow.extend(['New']*2)
    cols.extend(['SMILES','molAlogP'])
    for atype in set(paramdict['Assay Types']):
        headerrow.extend([atype]*(len(assaycols[atype])))
        cols.extend(assaycols[atype])
    cols.extend(['L/D/N?','N-Alkyl?','ID']*peplength)
    cols.extend(['Stereo Pattern'])
    cols.extend(['Methyl Pattern'])
    headerrow.extend(['New']*(3*peplength+2))
    mws.append(headerrow)
    mws.append(cols)
    #Match data
    #Experiment:Sequence Mass:Sequence Time:Assay:Assay Mass:Assay Time
    #Sometimes there are multiple matches within the Rt threshold to a single sequencing run peak due to the time imprecision of MS2-based Rt estimation. In that situation, the best guess would be the closest Rt to matching.
    matchdict = collections.defaultdict(lambda : collections.defaultdict(lambda : collections.defaultdict(lambda : collections.defaultdict(list))))
    for atype in set(paramdict['Assay Types']):
        for i,expt in enumerate(paramdict['Experiment Names']):
            try:
                assay_masses = sorted(assaydata[atype][expt].keys())
                sequence_masses = sorted(seqdata[expt].keys())
            except KeyError:
                print("Something wrong in experiment declaration vs actual data, key {}.".format(expt))
                sys.exit(1)
            try:
                test = sequence_masses[0] + 0.2
            except TypeError:
                print("Sequence data is formatted incorrectly")
                sys.exit(1)
            mass_prec = paramdict['Mass Precision (m/z)']
            mass_offset = paramdict['Assay Data Mass Offset'][i]
            time_prec = paramdict['Time Precision (s)']
            time_offset = paramdict['Assay Data Rt Offset'][i]
            for amass,smass in quickintervalcompare(assay_masses,sequence_masses,mass_prec,mass_offset,True):
                assay_times = sorted(assaydata[atype][expt][amass].keys())
                sequence_times = sorted(seqdata[expt][smass])
                for atime, stime in quickintervalcompare(assay_times,sequence_times,time_prec,time_offset,False):
                    #Record matches
                    matchdict[expt][smass][stime][atype].append([expt,names[atype][expt][amass][atime],amass,atime])
                    matchdict[expt][smass][stime][atype].append(assaydata[atype][expt][amass][atime])
                #Determine best match by proximity if there are multiple assay peaks near the MS2 data.
                for stime in matchdict[expt][smass].keys():
                    matchlist = matchdict[expt][smass][stime][atype]
                    if len(matchlist) > 2:
                        bestdistance = time_prec+1
                        bestindex = 0
                        for matchidx in range(0,len(matchlist),2):
                            atime = matchlist[matchidx][3]
                            dist = abs(stime-atime)
                            if dist < bestdistance:
                                bestdistance = dist
                                bestindex = matchidx
                        matchdict[expt][smass][stime][atype] = [matchlist[bestindex],matchlist[bestindex+1]]       
    #Write results
    for expt in matchdict.keys():
        for smass in matchdict[expt].keys():
            for stime in matchdict[expt][smass].keys():
                tmprow = []
                assaydata_towrite={}
                firstmatchedassay=True
                for i,atype in enumerate(set(paramdict['Assay Types'])):
                    if matchdict[expt][smass][stime][atype] == []: #Fill blanks if missing a match on some assays such that columns all still line up.
                        assaydata_towrite[atype]=[None]*len(assaycols[atype])
                        continue
                    if firstmatchedassay:
                        firstmatchedassay=False
                        tmprow.extend(matchdict[expt][smass][stime][atype][0])#Universal assay fields ['Expt','Name','Mass', 'Rt (s)']
                        tmprow.extend(seqdata[expt][smass][stime]) #Sequencing data
                        #Generate SMILES and molAlogP
                        compoundname = seqdata[expt][smass][stime][0]
                        compoundlist = compoundname.split(',')
                        SMILESlist = (compoundname.replace('*','')).split(',')
                        badname,smiles=make_SMILES(SMILESlist,*SMILESparams)
                        mol = Chem.MolFromSmiles(smiles)
                        mol = Chem.AddHs(mol)
                        molAlogP = Crippen.MolLogP(mol)
                        tmprow.extend([smiles,molAlogP])
                        #build residue analysis columns.
                        residueanalysiscols = []
                        
                        stereocol = []
                        methylcol = []
                        for x in range(peplength): 
                            residueanalysiscols.extend([aa_prop_dict[SMILESlist[x]]['Stereo'],aa_prop_dict[SMILESlist[x]]['N-alkyl'],compoundlist[x]])
                            
                            stereocol.append(aa_prop_dict[SMILESlist[x]]['Stereo'])
                            
                            if aa_prop_dict[SMILESlist[x]]['N-alkyl'] == True:
                                methylcol.append("Y")
                            else:
                                methylcol.append("N")
                        residueanalysiscols.append(fill_cell("".join(stereocol), cell_stereo))
                        residueanalysiscols.append(fill_cell("".join(methylcol), cell_methyl))
                        
                    data = matchdict[expt][smass][stime][atype][1]
                    assaydata_towrite[atype] = data
                for atype in set(paramdict['Assay Types']):
                    tmprow.extend(assaydata_towrite[atype])
                tmprow.extend(residueanalysiscols) #add residue analysis columns.
                for i,element in enumerate(tmprow):
                    if str(element).startswith('=HYPERLINK'):
                        tmprow[i] = fill_cell(str(element), copy.copy(cell_link))
                        
                mws.append(tmprow)
    if options.outfile != '':
        mwb.save('{}_Merged.xlsx'.format(options.outfile))
    else:
        mwb.save('Merged.xlsx')

if __name__ == '__main__':
    main()